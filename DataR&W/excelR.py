import openpyxl        # 导入openpyxl库，用于读取Excel文件


def createDataSet1(path):
    # 从Excel文件中读取数据并生成数据集和标签
    workbook = openpyxl.load_workbook(path)  # 打开Excel文件
    sheet = workbook.active  # 获取活动工作表
    dataSetX = []  # 初始化数据集
    dataSety = []  # 初始化数据集
    labels = [cell.value for cell in sheet[1][1:]]  # 获取特征标签（第一行的值，去掉第一列）

    for row in sheet.iter_rows(min_row=2):
        dataSetX.append([cell.value for cell in row[1:3]])  # 逐行读取数据并加入数据集
        dataSety.append([cell.value for cell in row[3:]])
        print(row)

    return dataSetX, dataSety  # 返回数据集和标签


def createDataSet2(path):
    # 从Excel文件中读取数据并生成数据集和标签
    workbook = openpyxl.load_workbook(path)  # 打开Excel文件
    sheet = workbook.active  # 获取活动工作表
    dataSetX = []  # 初始化数据集
    dataSety = []  # 初始化数据集
    data = []  # 初始化二维数组，用于存储数据集
    # labels = [cell.value for cell in sheet[1][1:]]  # 获取特征标签（第一行的值，去掉第一列）

    # 初始化当前列数据的临时数组
    current_column_data = []

    # 以下代码为读取Excel文件并处理每一行的逻辑
    for index, cell in enumerate(sheet.iter_rows(min_row=2, max_row=sheet.max_row, min_col=3, max_col=3, values_only=True), start=1):
        # 将单元格的值添加到临时数组
        current_column_data.append(cell[0])
        print(index)

    # 每 36 行结束时，将当前列数据添加到二维数组，并重置临时数组
        if index % 36 == 0:
            data.append(current_column_data)
            current_column_data = []
            print(index)

    return data, dataSety  # 返回数据集和标签


path = '2023C3.xlsx'  # Excel文件路径
X, y = createDataSet2(path=path)  # 读取数据集和标签

# 输出部分数据以供查看
print("Sample data points:")
print(X[:10])  # 输出前10个数据点
print("Sample labels:")
print(y[:10])  # 输出前10个标签
